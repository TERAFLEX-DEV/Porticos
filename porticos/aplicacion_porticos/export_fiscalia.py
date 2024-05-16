import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

class ExportarFiscalia:
    def __init__(self, listas_datos):
        self.listas_datos = listas_datos

    def exportar_excel(self, nombre_archivo):
        # Crear un nuevo libro de Excel
        libro = openpyxl.Workbook()

        # Escribir cada conjunto de datos en una hoja de Excel separada
        for idx, datos in enumerate(self.listas_datos, start=1):
            hoja = libro.create_sheet(title=f'Hoja {idx}')
            self._escribir_datos_en_hoja(hoja, datos)

        # Eliminar la hoja inicial predeterminada
        libro.remove(libro['Sheet'])

        # Guardar el archivo Excel
        libro.save(nombre_archivo)

    def _escribir_datos_en_hoja(self, hoja, datos):
        if isinstance(datos, list) and datos and isinstance(datos[0], dict):
            # Escribir los encabezados de las columnas
            encabezados = datos[0].keys()
            for idx, encabezado in enumerate(encabezados, start=1):
                hoja.cell(row=1, column=idx, value=encabezado)

            # Escribir los datos en las filas siguientes
            for fila_idx, fila_datos in enumerate(datos, start=2):
                for col_idx, encabezado in enumerate(encabezados, start=1):
                    hoja.cell(row=fila_idx, column=col_idx, value=fila_datos[encabezado])
        elif isinstance(datos, dict):
            # Escribir los datos del diccionario en dos columnas
            for fila_idx, (clave, valor) in enumerate(datos.items(), start=1):
                hoja.cell(row=fila_idx, column=1, value=clave)
                hoja.cell(row=fila_idx, column=2, value=valor)
        else:
            raise ValueError("El formato de los datos no es compatible.")