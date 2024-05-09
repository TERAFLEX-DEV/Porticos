import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

class Exportar:
    def __init__(self, listas_datos, ciudad, inicio, fin):
        self.listas_datos = listas_datos
        self.ciudad = ciudad
        self.inicio = inicio if isinstance(inicio, datetime) else datetime.strptime(inicio, "%Y-%m-%d %H:%M:%S")
        self.fin = fin if isinstance(fin, datetime) else datetime.strptime(fin, "%Y-%m-%d %H:%M:%S")
        self.inicio = self.inicio.strftime("%d-%m-%Y")
        self.fin = self.fin.strftime("%d-%m-%Y")

    def exportar_excel(self, nombre_archivo):
        # Crear un nuevo libro de Excel
        libro = openpyxl.Workbook()

        # Escribir cada arreglo de datos en una hoja de Excel separada
        for idx, datos in enumerate(self.listas_datos, start=1):
            hoja = libro.create_sheet(title=f'Hoja {idx}')
            self._escribir_datos_en_hoja(hoja, datos)

        # Eliminar la hoja inicial predeterminada
        libro.remove(libro['Sheet'])

        # Guardar el archivo Excel
        libro.save(nombre_archivo)

    def _escribir_datos_en_hoja(self, hoja, datos):
        hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)  
        hoja.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
        hoja.cell(row=1, column=1, value=f'Está imprimiendo {self.ciudad}')
        hoja.cell(row=2, column=1, value=f'En el rango de fecha {self.inicio} y {self.fin}')
    
        for fila, clave in enumerate(datos, start=3):
            valor = datos[clave]  # Obtener el valor correspondiente a la clave
            hoja.cell(row=fila, column=1, value=clave)  # Escribir la clave en la primera columna
            hoja.cell(row=fila, column=2, value=valor)  # Escribir el valor en la segunda columna